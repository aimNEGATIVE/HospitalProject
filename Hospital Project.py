from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import json
from openpyxl import load_workbook
import datetime

win = Tk()
win.title("")
win.geometry("900x500+500+200")
win.configure(bg="#FFFFFF")
win.resizable(False, False)
win.iconbitmap("images/hospital.ico")

clear_boxex_img = PhotoImage(file="buttons/clear-format.png")
background_frame_img = PhotoImage(file="images/background.png")
adminDict = {"admin": "admin"}

global doctor_dataDict

doctor_dataDict = {}
d = json.load(open("doctors.txt"))
doctor_dataDict.update(d)

################################################################################################################################
########################################################## Login Page ##########################################################
################################################################################################################################

login_frame = Frame(win, width=900, height=500, bg="#FFFFFF")
login_frame.place(x=0, y=0)

staff_img = PhotoImage(file="images/medical_personnel.png")
Label(login_frame, image=staff_img, bg="#FFFFFF").place(x=50, y=50)

title_lbl = Label(login_frame, text="Personnel Login", fg="#58C3EB", bg="#FFFFFF", font=("Microsoft YaHei UI", 25, "bold"))
title_lbl.place(x=550,y=80)

##### Username entry #####

username_entry = Entry(login_frame, width=30, bd=0, bg="#FFFFFF", font=("Microsoft YaHei UI", 13))
username_entry.place(x=540, y=150)
username_entry.insert(0, "Username")

def Enter_username_entry(event):     # entry-də "Username" yazısı olur. İstfadəci üzərinə klik etdikdə yazı yox olur və daxil etmək olur.
    username_entry.delete(0, "end")

def Leave_username_entry(event):
    username = username_entry.get()
    if username == "":
        username_entry.insert(0, "Username")

username_entry.bind("<FocusIn>", Enter_username_entry)
username_entry.bind("<FocusOut>", Leave_username_entry)

Frame(login_frame, width=310, height=2, bg="black").place(x=535,y=177)

##### Password entry #####

password_entry = Entry(login_frame, width=30, bd=0, bg="#FFFFFF", font=("Microsoft YaHei UI", 13))
password_entry.place(x=540, y=220)
password_entry.insert(0, "Password")

def Enter_password_entry(event):  # entry-də "Password" yazısı olur. İstfadəci üzərinə klik etdikdə yazı yox olur və daxil etmək olur.
    password_entry.config(show="•")
    password_entry.delete(0, "end")

def Leave_password_entry(event):
    password = password_entry.get()
    if password == "":
        password_entry.config(show="")
        password_entry.insert(0, "Password")

password_entry.bind("<FocusIn>", Enter_password_entry)
password_entry.bind("<FocusOut>", Leave_password_entry)

Frame(login_frame, width=310, height=2, bg="black").place(x=535,y=247)

##### Show/hide password button #####

def ShowHidePassword():     # Password-u göstərmək və ya gizlətmək funksiyası
    if password_entry.cget("show") == "•":
        password_entry.config(show="")
        showhide_btn.config(image=hide_img)
    else:
        password_entry.config(show="•")
        showhide_btn.config(image=show_img)

show_img = PhotoImage(file="buttons/eye.png")
hide_img = PhotoImage(file="buttons/eye-crossed.png")
showhide_btn = Button(login_frame, image=show_img, bg="#FFFFFF", bd=0, activebackground="#FFFFFF", cursor="hand2", command=ShowHidePassword)
showhide_btn.place(x=810, y=215)

##### Login button #####

def LoginButton():
    if len(username_entry.get()) == 0 or len(password_entry.get()) == 0 or username_entry.get() == "Username" or password_entry.get() == "Password":
        messagebox.showerror(title="Login", message="All required fields must be filled out!")
    elif username_entry.get() in adminDict and adminDict[username_entry.get()] == password_entry.get():
        win.state("zoomed")
        login_frame.place_forget()
        app_frame.place(x=0, y=0)
    else:
        messagebox.showerror(title="Login", message="The username or password is incorrect!")
    
login_btn = Button(login_frame, text="Login", font=("Microsoft YaHei UI", 13), fg="#FFFFFF", width=27, bd=0, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=LoginButton)
login_btn.place(x=555, y=300)

################################################################################################################################
########################################################## App frame ###########################################################
################################################################################################################################

app_frame = Frame(win, width=1920, height=1080, bg="#C6EEFF") #place on Login button
head_line = Frame(app_frame, width=1920, height=70, bg="#79D4F6")
head_line.place(x=0, y=0)

Label(app_frame, image=background_frame_img, bd=0).place(x=350, y=100)

logo_img = PhotoImage(file="images/hospital.png")
logo_lbl = Label(head_line, image=logo_img, bg="#79D4F6")
logo_lbl.place(x=600, y=0)

Label(head_line, text="Central Medical Center", font=("Microsoft YaHei UI", 35, "bold"), bg="#79D4F6").place(x=700, y=2)

##### Side menu #####

def SideMenuButton():       # yan panel frame-in aktivləşməsi
    if side_menu_frame.cget("width") == 0:
        side_menu_frame.config(width=300)
    else:
        side_menu_frame.config(width=0)

side_menu_img = PhotoImage(file="buttons/menu.png")
side_menu_btn = Button(app_frame, image=side_menu_img, cursor="hand2", bg="#79D4F6", bd=0, activebackground="#FFFFFF", command=SideMenuButton)
side_menu_btn.place(x=2, y=2)

side_menu_frame = Frame(app_frame, width=0, height=1000, bg="#79D4F6")
side_menu_frame.place(x=0, y=70)

##### DOCTORS EDIT #####

def DoctorsEdit():
    global search_entry, search_top, doctor_dataDict

    edit_doctors_top = Toplevel()
    edit_doctors_top.title("Edit doctors")
    edit_doctors_top.iconbitmap("images/hospital.ico")
    edit_doctors_top.geometry("1000x500+450+200")
    edit_doctors_top.resizable(False, False)
    edit_doctors_top.grab_set()

    def Search():
        global search_entry, search_top

        search_top = Toplevel(edit_doctors_top)
        search_top.title("Search Data")
        search_top.geometry("400x200+750+350")
        search_top.resizable(False, False)
        search_top.iconbitmap("images/hospital.ico")

        search_frame = LabelFrame(search_top, text="Surname")
        search_frame.pack(padx=10, pady=10)

        search_entry = Entry(search_frame, font=("Microsoft YaHei UI", 15))
        search_entry.pack(padx=20, pady=20)

        search_btn = Button(search_top, text="Search", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=SearchData)
        search_btn.pack(padx=20, pady=20)

    def ResetSearch():
        for data in sheet_tree.get_children():
            sheet_tree.delete(data)

        global doctor_count
        for FIN in doctor_dataDict:
            if doctor_count % 2 == 0:
                sheet_tree.insert(parent='', index='end', iid=doctor_count, text='', values=(FIN, doctor_dataDict[FIN][0], doctor_dataDict[FIN][1], doctor_dataDict[FIN][2], doctor_dataDict[FIN][3], doctor_dataDict[FIN][4]), tags=("evenrow",))
            else:
                sheet_tree.insert(parent='', index='end', iid=doctor_count, text='', values=(FIN, doctor_dataDict[FIN][0], doctor_dataDict[FIN][1], doctor_dataDict[FIN][2], doctor_dataDict[FIN][3], doctor_dataDict[FIN][4]), tags=("oddrow",))
        
            doctor_count += 1
   
    def SearchData():
        check = 0

        for i in doctor_dataDict:
            if search_entry.get() == doctor_dataDict[i][1]:
                check += 1
            else:
                pass

        if check != 0:
            for data in sheet_tree.get_children():
                sheet_tree.delete(data)

            for k in doctor_dataDict:
                if search_entry.get() == doctor_dataDict[k][1]:
                    global doctor_count  
                    if doctor_count % 2 == 0:
                        sheet_tree.insert(parent='', index='end', iid=doctor_count, text='', values=(k, doctor_dataDict[k][0], doctor_dataDict[k][1], doctor_dataDict[k][2], doctor_dataDict[k][3], doctor_dataDict[k][4]), tags=("evenrow",))
                    else:
                        sheet_tree.insert(parent='', index='end', iid=doctor_count, text='', values=(k, doctor_dataDict[k][0], doctor_dataDict[k][1], doctor_dataDict[k][2], doctor_dataDict[k][3], doctor_dataDict[k][4]), tags=("oddrow",))
            
                    doctor_count += 1
    
            search_top.destroy()
        else:
            messagebox.showerror(title="Search", message="This surname is not exist")  

    menu_main = Menu(edit_doctors_top)  # Axtarış üçün menyunun yaradılması
    edit_doctors_top.config(menu=menu_main)

    search_menu = Menu(menu_main, tearoff=0)
    menu_main.add_cascade(label="Searching", menu=search_menu)  
    

    search_menu.add_command(label="Search", command=Search)       # axtarış menyusunun daxilində olan komandalar "Search" və "Reset"
    search_menu.add_command(label="Reset", command=ResetSearch)


    ##### Sheet #####

    style = ttk.Style()   
    style.theme_use('default')
    style.configure("Treeview", background="#D3D3D3", foreground="black", rowheight=25, fieldbackground="#D3D3D3")
    style.map("Treeview", background=[("selected", "#79D4F6")])

    sheet_frame = Frame(edit_doctors_top)
    sheet_frame.pack(pady=10)

    sheet_scroll = ttk.Scrollbar(sheet_frame)  # Cədvəldə scrollbar yaradılması
    sheet_scroll.pack(side=RIGHT, fill=Y) # sağ tərəf və Y oxu istiqamətinə

    sheet_tree = ttk.Treeview(sheet_frame, yscrollcommand=sheet_scroll.set, selectmode="extended")
    sheet_tree.pack()

    sheet_scroll.config(command=sheet_tree.yview)

    sheet_tree["columns"] = ("FIN", "Name", "Surname", "Position", "Phone", "Cabinet")  # Kolonların adları

    sheet_tree.column("#0", width=0, stretch=NO)
    sheet_tree.column("FIN", anchor=CENTER, width=90)
    sheet_tree.column("Name", anchor=CENTER, width=140)
    sheet_tree.column("Surname", anchor=CENTER, width=140)            # Kolonların məlumatların yerləşimi və default eni
    sheet_tree.column("Position", anchor=CENTER, width=140)
    sheet_tree.column("Phone", anchor=CENTER, width=140)
    sheet_tree.column("Cabinet", anchor=CENTER, width=140)

    sheet_tree.heading("#0", text="", anchor=W)
    sheet_tree.heading("FIN", text="FIN", anchor=CENTER)
    sheet_tree.heading("Name", text="Name", anchor=CENTER)
    sheet_tree.heading("Surname", text="Surname", anchor=CENTER)      # Kolonların başlıqlarının yerləşimi və default eni
    sheet_tree.heading("Position", text="Position", anchor=CENTER)
    sheet_tree.heading("Phone", text="Phone", anchor=CENTER)
    sheet_tree.heading("Cabinet", text="Cabinet", anchor=CENTER)

    #### DICTIOANRY #####

    d = json.load(open("doctors.txt"))
    doctor_dataDict.update(d)

    sheet_tree.tag_configure("oddrow", background="white")        # Cədvəldəki cərgələrin rənglərinin təyin olunması
    sheet_tree.tag_configure("evenrow", background="lightblue")

#### Sheet inserting ####

    global doctor_count
    doctor_count = 0
   
    for FIN in doctor_dataDict:         # Cədvələ data əlavə edilməsi     
        if doctor_count % 2 == 0:
            sheet_tree.insert(parent='', index='end', iid=doctor_count, text='', values=(FIN, doctor_dataDict[FIN][0], doctor_dataDict[FIN][1], doctor_dataDict[FIN][2], doctor_dataDict[FIN][3], doctor_dataDict[FIN][4]), tags=("evenrow",))
        else:
            sheet_tree.insert(parent='', index='end', iid=doctor_count, text='', values=(FIN, doctor_dataDict[FIN][0], doctor_dataDict[FIN][1], doctor_dataDict[FIN][2], doctor_dataDict[FIN][3], doctor_dataDict[FIN][4]), tags=("oddrow",))
        
        doctor_count += 1

    ##### Labels and entries #####

    data_frame = Frame(edit_doctors_top, width=1000, height=219)
    data_frame.place(x=0, y=280)

    fin_lbl = Label(data_frame, text="FIN", font=("Microsoft YaHei UI", 13, "bold"))
    fin_lbl.place(x=50, y=20)
    fin_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    fin_entry.place(x=120, y=20)

    name_lbl = Label(data_frame, text="Name", font=("Microsoft YaHei UI", 13, "bold"))
    name_lbl.place(x=50, y=60)
    name_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    name_entry.place(x=120, y=60)

    surname_lbl = Label(data_frame, text="Surname",font=("Microsoft YaHei UI", 13, "bold"))
    surname_lbl.place(x=350, y=20)
    surname_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    surname_entry.place(x=450, y=20)

    position_lbl = Label(data_frame, text="Position",font=("Microsoft YaHei UI", 13, "bold"))
    position_lbl.place(x=350, y=60)
    position_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    position_entry.place(x=450, y=60)

    phone_lbl = Label(data_frame, text="Phone",font=("Microsoft YaHei UI", 13, "bold"))
    phone_lbl.place(x=680, y=20)
    phone_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    phone_entry.place(x=760, y=20)

    cabinet_lbl = Label(data_frame, text="Cabinet",font=("Microsoft YaHei UI", 13, "bold"))
    cabinet_lbl.place(x=680, y=60)
    cabinet_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    cabinet_entry.place(x=760, y=60)

    ##### Buttons #####

    ### Add Data ###

    def AddData():
        global doctor_count      
        if len(fin_entry.get()) == 0 or len(name_entry.get()) == 0 or len(surname_entry.get()) == 0  or len(position_entry.get()) == 0 or len(phone_entry.get()) == 0 or len(cabinet_entry.get()) == 0:
            messagebox.showwarning(title="Error", message="All required fields must be filled out!")
        elif fin_entry.get() in doctor_dataDict:
            messagebox.showerror(title="FIN error", message="Person with this FIN is already in database!")
        elif len(fin_entry.get()) != 7:
            messagebox.showerror(title="FIN error", message="FIN must contain 7 characters!")
        else:
            choise = messagebox.askquestion(title="Add data", message="Do you want to add new data?")
            if choise == "yes":
                messagebox.showinfo(title="Add data", message="New data added successfully")
                newList = []
                values = [name_entry.get(), surname_entry.get(), position_entry.get(), phone_entry.get(), cabinet_entry.get()]
                for data in values:
                    newList.append(data)
                    doctor_dataDict.update({fin_entry.get(): newList})
                    json.dump(doctor_dataDict, open("doctors.txt", "w"))
                
                for data in sheet_tree.get_children():
                    sheet_tree.delete(data)

                for FIN in doctor_dataDict:
                    if doctor_count % 2 == 0:
                        sheet_tree.insert(parent='', index='end', iid=doctor_count, text='', values=(FIN, doctor_dataDict[FIN][0], doctor_dataDict[FIN][1], doctor_dataDict[FIN][2], doctor_dataDict[FIN][3], doctor_dataDict[FIN][4]), tags=("evenrow",))
                    else:
                        sheet_tree.insert(parent='', index='end', iid=doctor_count, text='', values=(FIN, doctor_dataDict[FIN][0], doctor_dataDict[FIN][1], doctor_dataDict[FIN][2], doctor_dataDict[FIN][3], doctor_dataDict[FIN][4]), tags=("oddrow",))
        
                    doctor_count += 1
            
    ### Remove Data ###

    def RemoveOneSelected():
        if len(fin_entry.get()) == 0 and len(name_entry.get()) == 0 and len(surname_entry.get()) == 0 and len(position_entry.get()) == 0 and len(phone_entry.get()) == 0 and len(cabinet_entry.get()) == 0:
            messagebox.showinfo(title="Removing data", message="First select the cell to delete")
        else:
            delete_choise = messagebox.askquestion(title="Deleting data", message="Are you sure you want to delete this item from list?") 
            if delete_choise == "yes":
                selected_data = sheet_tree.focus()
                value = sheet_tree.item(selected_data, "values") # value[0]  fin code 
                x = sheet_tree.selection()[0]
                sheet_tree.delete(x)
                doctor_dataDict.pop(value[0])
                json.dump(doctor_dataDict, open("doctors.txt", "w"))
                messagebox.showinfo(title="Deleting data", message="Data has been successfully deleted")

                fin_entry.delete(0, END)
                name_entry.delete(0, END)
                surname_entry.delete(0, END)
                position_entry.delete(0, END)
                phone_entry.delete(0, END)
                cabinet_entry.delete(0, END)

    ### Update Data ###

    def UpdateData():
        if len(fin_entry.get()) == 0 and len(name_entry.get()) == 0 and len(surname_entry.get()) == 0 and len(position_entry.get()) == 0 and len(phone_entry.get()) == 0 and len(cabinet_entry.get()) == 0:
            messagebox.showinfo(title="Removing data", message="First select the cell to update")
        else:
            selected_data = sheet_tree.focus()
            value = sheet_tree.item(selected_data, "values")
            if fin_entry.get() != value[0]:  # ferqlidirse  error vereecekl
                messagebox.showerror(title="Error", message="FIN cannot be changed!")
                fin_entry.delete(0, END)
                fin_entry.insert(0, value[0])
            elif name_entry.get() == value[1] and surname_entry.get() == value[2] and position_entry.get() == value[3] and phone_entry.get() == value[4] and cabinet_entry.get() == value[5]:
                messagebox.showerror(title="Update Error", message="No changes detected!")
            else:
                choise = messagebox.askquestion(title="Updating", message="Are you sure you want to update data?")
                if choise == "yes":
                    sheet_tree.item(selected_data, text="", values=(fin_entry.get(), name_entry.get(), surname_entry.get(), position_entry.get(), phone_entry.get(), cabinet_entry.get()))
                    values=(name_entry.get(), surname_entry.get(), position_entry.get(), phone_entry.get(), cabinet_entry.get())

                    newdataList = []
                    for data in values:
                        newdataList.append(data) 

                    doctor_dataDict.update({fin_entry.get(): newdataList})
                    json.dump(doctor_dataDict, open("doctors.txt", "w"))

                    fin_entry.delete(0, END)
                    name_entry.delete(0, END)
                    surname_entry.delete(0, END)
                    position_entry.delete(0, END)
                    phone_entry.delete(0, END)
                    cabinet_entry.delete(0, END)

                    messagebox.showinfo(title="Data updates", message="Data was succsessfully updated" )

    ### Clear Entries ###

    def ClearEntries():
        fin_entry.delete(0, END)
        name_entry.delete(0, END)
        surname_entry.delete(0, END)
        position_entry.delete(0, END)
        phone_entry.delete(0, END)
        cabinet_entry.delete(0, END)

    ### Select Data ###

    def SelectData(event):   # Cərgəyə klik etdikdə, orada olan datanın entry-lərə ötürülməsi
        fin_entry.delete(0, END)
        name_entry.delete(0, END)
        surname_entry.delete(0, END)
        position_entry.delete(0, END)
        phone_entry.delete(0, END)
        cabinet_entry.delete(0, END)

        selected_data = sheet_tree.focus()
        values = sheet_tree.item(selected_data, "values")

        fin_entry.insert(0, values[0])
        name_entry.insert(0, values[1])
        surname_entry.insert(0, values[2])
        position_entry.insert(0, values[3])
        phone_entry.insert(0, values[4])
        cabinet_entry.insert(0, values[5])

    update_data_btn = Button(data_frame, text="Update Data", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=UpdateData)
    update_data_btn.place(x=50, y=150)

    add_data_btn = Button(data_frame, text="Add Data", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=AddData)
    add_data_btn.place(x=170, y=150)

    remove_one_btn = Button(data_frame, text="Remove Selected", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=RemoveOneSelected)
    remove_one_btn.place(x=265, y=150)

    clear_boxes_btn = Button(data_frame, image=clear_boxex_img, fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", height=26, width=26, command=ClearEntries)
    clear_boxes_btn.place(x=900, y=150)

    sheet_tree.bind("<ButtonRelease-1>", SelectData)


###################################### PATIENTS EDIT ######################################

def PatientsEdit():
    global search_entry2, search_top2
    
    edit_patients_top = Toplevel()
    edit_patients_top.title("Edit patients")
    edit_patients_top.iconbitmap("images/hospital.ico")
    edit_patients_top.geometry("1000x500+450+200")
    edit_patients_top.resizable(False, False)
    edit_patients_top.grab_set()

    def Search():
        global search_entry2, search_top2

        search_top2 = Toplevel(edit_patients_top)
        search_top2.title("Search Data")
        search_top2.geometry("400x200+750+350")
        search_top2.resizable(False, False)
        search_top2.iconbitmap("images/hospital.ico")

        search_frame = LabelFrame(search_top2, text="Surname")
        search_frame.pack(padx=10, pady=10)

        search_entry2 = Entry(search_frame, font=("Microsoft YaHei UI", 15))
        search_entry2.pack(padx=20, pady=20)

        search_btn = Button(search_top2, text="Search", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=SearchData)
        search_btn.pack(padx=20, pady=20)

    def ResetSearch():
        for data in sheet_tree.get_children():
            sheet_tree.delete(data)

        global patient_count
        for FIN in patient_dataDict:
            if patient_count % 2 == 0:
                sheet_tree.insert(parent='', index='end', iid=patient_count, text='', values=(FIN, patient_dataDict[FIN][0], patient_dataDict[FIN][1], patient_dataDict[FIN][2], patient_dataDict[FIN][3], patient_dataDict[FIN][4], patient_dataDict[FIN][5]), tags=("evenrow",))
            else:
                sheet_tree.insert(parent='', index='end', iid=patient_count, text='', values=(FIN, patient_dataDict[FIN][0], patient_dataDict[FIN][1], patient_dataDict[FIN][2], patient_dataDict[FIN][3], patient_dataDict[FIN][4], patient_dataDict[FIN][5]), tags=("oddrow",))

            patient_count += 1

    def SearchData():
        check = 0

        for i in patient_dataDict:
            if search_entry2.get() == patient_dataDict[i][1]:
                check +=1
            else:
                pass
       
        if check !=0:
            for data in sheet_tree.get_children():
                sheet_tree.delete(data)

            for k in patient_dataDict:
                if search_entry2.get() == patient_dataDict[k][1]:
                    global patient_count
                    if patient_count % 2 == 0:
                        sheet_tree.insert(parent='', index='end', iid=patient_count, text='', values=(k, patient_dataDict[k][0], patient_dataDict[k][1], patient_dataDict[k][2], patient_dataDict[k][3], patient_dataDict[k][4], patient_dataDict[k][5]), tags=("evenrow",))
                    else:
                        sheet_tree.insert(parent='', index='end', iid=patient_count, text='', values=(k, patient_dataDict[k][0], patient_dataDict[k][1], patient_dataDict[k][2], patient_dataDict[k][3], patient_dataDict[k][4], patient_dataDict[k][5]), tags=("oddrow",))
        
                    patient_count += 1

            search_top2.destroy()
        else:
            messagebox.showerror(title="Search", message="This surname is not exist")  

    menu_main = Menu(edit_patients_top)
    edit_patients_top.config(menu=menu_main)

    search_menu = Menu(menu_main, tearoff=0)
    menu_main.add_cascade(label="Searching", menu=search_menu)
    
    search_menu.add_command(label="Search", command=Search)
    search_menu.add_command(label="Reset", command=ResetSearch)

    ##### Sheet #####

    style = ttk.Style()
    style.theme_use('default')
    style.configure("Treeview", background="#D3D3D3", foreground="black", rowheight=25, fieldbackground="#D3D3D3")
    style.map("Treeview", background=[("selected", "#79D4F6")])

    sheet_frame = Frame(edit_patients_top)
    sheet_frame.pack(pady=10)

    sheet_scroll = ttk.Scrollbar(sheet_frame)
    sheet_scroll.pack(side=RIGHT, fill=Y)

    sheet_tree = ttk.Treeview(sheet_frame, yscrollcommand=sheet_scroll.set, selectmode="extended")
    sheet_tree.pack()

    sheet_scroll.config(command=sheet_tree.yview)

    sheet_tree["columns"] = ("FIN", "Name", "Surname", "Phone", "Address", "Age", "Blood Type")

    sheet_tree.column("#0", width=0, stretch=NO)
    sheet_tree.column("FIN", anchor=CENTER, width=90)
    sheet_tree.column("Name", anchor=CENTER, width=140)
    sheet_tree.column("Surname", anchor=CENTER, width=140)
    sheet_tree.column("Phone", anchor=CENTER, width=140)
    sheet_tree.column("Address", anchor=CENTER, width=140)
    sheet_tree.column("Age", anchor=CENTER, width=140)
    sheet_tree.column("Blood Type", anchor=CENTER, width=140)

    sheet_tree.heading("#0", text="", anchor=W)
    sheet_tree.heading("FIN", text="FIN", anchor=CENTER)
    sheet_tree.heading("Name", text="Name", anchor=CENTER)
    sheet_tree.heading("Surname", text="Surname", anchor=CENTER)
    sheet_tree.heading("Phone", text="Phone", anchor=CENTER)
    sheet_tree.heading("Address", text="Address", anchor=CENTER)
    sheet_tree.heading("Age", text="Age", anchor=CENTER)
    sheet_tree.heading("Blood Type", text="Blood Type", anchor=CENTER)

    #### PATIENTS DICTIONARY ####
    global patient_dataDict
    patient_dataDict = {}
    p = json.load(open("patients.txt"))
    patient_dataDict.update(p)

    sheet_tree.tag_configure("oddrow", background="white")
    sheet_tree.tag_configure("evenrow", background="lightblue")

    #### Sheet inserting ####

    global patient_count
    patient_count = 0

    for FIN in patient_dataDict:
        if patient_count % 2 == 0:
            sheet_tree.insert(parent='', index='end', iid=patient_count, text='', values=(FIN, patient_dataDict[FIN][0], patient_dataDict[FIN][1], patient_dataDict[FIN][2], patient_dataDict[FIN][3], patient_dataDict[FIN][4], patient_dataDict[FIN][5]), tags=("evenrow",))
        else:
            sheet_tree.insert(parent='', index='end', iid=patient_count, text='', values=(FIN, patient_dataDict[FIN][0], patient_dataDict[FIN][1], patient_dataDict[FIN][2], patient_dataDict[FIN][3], patient_dataDict[FIN][4], patient_dataDict[FIN][5]), tags=("oddrow",))
        
        patient_count += 1

    ##### Labels and entries #####

    data_frame = Frame(edit_patients_top, width=1000, height=219)
    data_frame.place(x=0, y=280)

    fin_lbl = Label(data_frame, text="FIN", font=("Microsoft YaHei UI", 13, "bold"))
    fin_lbl.place(x=50, y=20)
    fin_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    fin_entry.place(x=120, y=20)

    name_lbl = Label(data_frame, text="Name", font=("Microsoft YaHei UI", 13, "bold"))
    name_lbl.place(x=50, y=60)
    name_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    name_entry.place(x=120, y=60)

    surname_lbl = Label(data_frame, text="Surname",font=("Microsoft YaHei UI", 13, "bold"))
    surname_lbl.place(x=350, y=20)
    surname_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    surname_entry.place(x=450, y=20)

    phone_lbl = Label(data_frame, text="Phone",font=("Microsoft YaHei UI", 13, "bold"))
    phone_lbl.place(x=350, y=60)
    phone_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    phone_entry.place(x=450, y=60)

    address_lbl = Label(data_frame, text="Address",font=("Microsoft YaHei UI", 13, "bold"))
    address_lbl.place(x=680, y=20)
    address_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    address_entry.place(x=760, y=20)

    age_lbl = Label(data_frame, text="Age",font=("Microsoft YaHei UI", 13, "bold"))
    age_lbl.place(x=680, y=60)
    age_entry = Entry(data_frame, font=("Microsoft YaHei UI", 13))
    age_entry.place(x=760, y=60)

    blood_lbl = Label(data_frame, text="Blood",font=("Microsoft YaHei UI", 13, "bold"))
    blood_lbl.place(x=680, y=100)
    blood_list = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
    blood_combo = ttk.Combobox(data_frame, values=blood_list, state="readonly", font=("Microsoft YaHei UI", 13, "bold"), width=4)
    blood_combo.place(x=760, y=100)
    blood_combo.current(0)   #### duzeltmek

    ##### Buttons #####

    ### Add Data ###

    def AddData():
        global patient_count
        if len(fin_entry.get()) == 0 or len(name_entry.get()) == 0 or len(surname_entry.get()) == 0 or len(phone_entry.get()) == 0 or len(address_entry.get()) == 0 or len(age_entry.get()) == 0:
            messagebox.showwarning(title="Error", message="All required fields must be filled out!")
        elif fin_entry.get() in patient_dataDict:
            messagebox.showerror(title="FIN error", message="Person with this FIN is already in database!")
        elif len(fin_entry.get()) != 7:
            messagebox.showerror(title="FIN error", message="FIN must contain 7 characters!")
        else:
            choise = messagebox.askquestion(title="Add data", message="Do you want to add new data?")
            if choise == "yes":
                messagebox.showinfo(title="Add data", message="New data added successfully")
                newList = []
                values = [name_entry.get(), surname_entry.get(), phone_entry.get(), address_entry.get(), age_entry.get(), blood_combo.get()]
                for data in values:
                    newList.append(data)
                    patient_dataDict.update({fin_entry.get(): newList})
                    json.dump(patient_dataDict, open("patients.txt", "w"))

                for data in sheet_tree.get_children():
                    sheet_tree.delete(data)

                for FIN in patient_dataDict:
                    if patient_count % 2 == 0:
                        sheet_tree.insert(parent='', index='end', iid=patient_count, text='', values=(FIN, patient_dataDict[FIN][0], patient_dataDict[FIN][1], patient_dataDict[FIN][2], patient_dataDict[FIN][3], patient_dataDict[FIN][4], patient_dataDict[FIN][5]), tags=("evenrow",))
                    else:
                        sheet_tree.insert(parent='', index='end', iid=patient_count, text='', values=(FIN, patient_dataDict[FIN][0], patient_dataDict[FIN][1], patient_dataDict[FIN][2], patient_dataDict[FIN][3], patient_dataDict[FIN][4], patient_dataDict[FIN][5]), tags=("oddrow",))
        
                    patient_count += 1
                
    ### Remove data ###

    def RemoveOneSelected():
        if len(fin_entry.get()) == 0 and len(name_entry.get()) == 0 and len(surname_entry.get()) == 0 and len(phone_entry.get()) == 0 and len(address_entry.get()) == 0 and len(age_entry.get()) == 0:
            messagebox.showinfo(title="Removing data", message="First select the cell to delete")
        else:
            delete_choise = messagebox.askquestion(title="Deleting data", message="Are you sure you want to delete this item from list?")
            if delete_choise == "yes":
                selected_data = sheet_tree.focus()
                value = sheet_tree.item(selected_data, "values")
                x = sheet_tree.selection()[0]
                sheet_tree.delete(x)
                patient_dataDict.pop(value[0])
                json.dump(patient_dataDict, open("patients.txt", "w"))
                messagebox.showinfo(title="Deleting data", message="Data has been successfully deleted")

                fin_entry.delete(0, END)
                name_entry.delete(0, END)
                surname_entry.delete(0, END)
                phone_entry.delete(0, END)
                address_entry.delete(0, END)
                age_entry.delete(0, END)
                blood_combo.current(0)

    ### Update Data ###

    def UpdateData(): 
        if len(fin_entry.get()) == 0 and len(name_entry.get()) == 0 and len(surname_entry.get()) == 0 and len(phone_entry.get()) == 0 and len(address_entry.get()) == 0 and len(age_entry.get()) == 0:
            messagebox.showinfo(title="Removing data", message="First select the cell to update")
        else:
            selected_data = sheet_tree.focus()
            value = sheet_tree.item(selected_data, "values")

            if fin_entry.get() != value[0]:
                messagebox.showerror(title="Error", message="FIN cannot be changed!")
                fin_entry.delete(0, END)
                fin_entry.insert(0, value[0])
            elif name_entry.get() == value[1] and surname_entry.get() == value[2] and phone_entry.get() == value[3] and address_entry.get() == value[4] and age_entry.get() == value[5] and blood_combo.get() == value[6]:
                messagebox.showerror(title="Update Error", message="No changes detected!")
            else:
                choise = messagebox.askquestion(title="Updating", message="Are you sure you want to update data?")
                if choise == "yes":
                    sheet_tree.item(selected_data, text="", values=(fin_entry.get(), name_entry.get(), surname_entry.get(), phone_entry.get(), address_entry.get(), age_entry.get(), blood_combo.get()))
                    values=(name_entry.get(), surname_entry.get(), phone_entry.get(), address_entry.get(), age_entry.get(), blood_combo.get())

                    newdataList = []
                    for data in values:
                        newdataList.append(data)

                    patient_dataDict.update({fin_entry.get(): newdataList})
                    json.dump(patient_dataDict, open("patients.txt", "w"))

                    fin_entry.delete(0, END)
                    name_entry.delete(0, END)
                    surname_entry.delete(0, END)
                    phone_entry.delete(0, END)
                    address_entry.delete(0, END)
                    blood_combo.current(0)

                    messagebox.showinfo(title="Data updates", message="Data was succsessfully updated")


    ### Clear Entries ###
    
    def ClearEntries():
        fin_entry.delete(0, END)
        name_entry.delete(0, END)
        surname_entry.delete(0, END)
        phone_entry.delete(0, END)
        address_entry.delete(0, END)
        age_entry.delete(0, END)
        blood_combo.current(0)

    ### Select Data ###

    def SelectData(event):
        blood_list = ["A+", "A-", "B+", "B-", "AB+", "AB-", "O+", "O-"]
        index=0

        fin_entry.delete(0, END)
        name_entry.delete(0, END)
        surname_entry.delete(0, END)
        phone_entry.delete(0, END)
        address_entry.delete(0, END)
        age_entry.delete(0, END)
        blood_combo.current(0)

        selected_data = sheet_tree.focus()
        values = sheet_tree.item(selected_data, "values")

        fin_entry.insert(0, values[0])
        name_entry.insert(0, values[1])
        surname_entry.insert(0, values[2])
        phone_entry.insert(0, values[3])
        address_entry.insert(0, values[4])
        age_entry.insert(0, values[5])
        for i in blood_list:           
            if i==values[6]:
                blood_combo.current(index)
            index+=1
                
    update_data_btn2 = Button(data_frame, text="Update Data", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=UpdateData)
    update_data_btn2.place(x=50, y=150)

    add_data_btn = Button(data_frame, text="Add Data", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=AddData)
    add_data_btn.place(x=170, y=150)

    remove_one_btn = Button(data_frame, text="Remove Selected", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=RemoveOneSelected)
    remove_one_btn.place(x=265, y=150)

    clear_boxes_btn = Button(data_frame, image=clear_boxex_img, fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", height=26, width=26, command=ClearEntries)
    clear_boxes_btn.place(x=900, y=150)

    sheet_tree.bind("<ButtonRelease-1>", SelectData)

    ############################ Appointment edit ############################

blood_test_img = PhotoImage(file="buttons/blood-test.png")

def AppointmentEdit():

    edit_appointment_top = Toplevel()
    edit_appointment_top.title("Appointment edit")
    edit_appointment_top.iconbitmap("images/hospital.ico")
    edit_appointment_top.geometry("1000x500+450+200")
    edit_appointment_top.resizable(False, False)
    edit_appointment_top.grab_set()

    ### Tests frame ###


    def BloodTest():  #duzeltmek

        global p_search_entry, p_search_top

        p_search_top = Toplevel(edit_appointment_top)
        p_search_top.title("Appointment")
        p_search_top.iconbitmap("images/hospital.ico")
        p_search_top.geometry("400x200+750+350")
        p_search_top.resizable(False, False)

        p_search_frame = LabelFrame(p_search_top, text="FIN")
        p_search_frame.pack(padx=10, pady=10)

        p_search_entry = Entry(p_search_frame, font=("Microsoft YaHei UI", 15))
        p_search_entry.pack(padx=20, pady=20)

        p_search_btn = Button(p_search_top, text="Search", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=PatientSearch_blood)
        p_search_btn.pack(padx=20, pady=20)

    def PatientSearch_blood():
        global fin
        patients_data = json.load(open("patients.txt"))
        fin = p_search_entry.get() 

        patients_blood_test_results = {}

        try:       # "Blood_test_results.txt" adlı faylın mövcudluğunun yoxlanılması
            results = json.load(open("blood_test_results.txt"))
        except IOError:
            if fin in patients_data:

                p_search_top.destroy()
                
                name = patients_data[fin][0]
                surname = patients_data[fin][1]
                age = patients_data[fin][4]

                #### Blood test ####

                blood_test_top = Toplevel(edit_appointment_top)
                blood_test_top.title("Blood test data")
                blood_test_top.geometry("600x800+700+150") 
                blood_test_top.resizable(False, False)           
                blood_test_top.grab_set()

                search_frame = LabelFrame(blood_test_top, text="Patient's name and surname", font=("Microsoft YaHei UI", 11, "bold"))
                search_frame.place(x=10, y=10)

                name_lbl = Label(search_frame, text=f"{name} {surname}", font=("Microsoft YaHei UI", 13))
                name_lbl.pack(padx=5, pady=5)

                birth_data_frame = LabelFrame(blood_test_top, text="Age", font=("Microsoft YaHei UI", 11, "bold"))
                birth_data_frame.place(x=500, y=10)

                birth_data_lbl = Label(birth_data_frame, text=age, font=("Microsoft YaHei UI", 13))
                birth_data_lbl.pack(padx=5, pady=5)

                appointed_doctor_frame = LabelFrame(blood_test_top, text="Appointed doctor", font=("Microsoft YaHei UI", 11, "bold")) 
                appointed_doctor_frame.place(x=280, y=10)

                app_doctors = []
                for i in doctor_dataDict:
                    app_doctors.append((doctor_dataDict[i][0] + " " + doctor_dataDict[i][1]))


                appointed_doctor_box = ttk.Combobox(appointed_doctor_frame, values=app_doctors, state="readonly")
                appointed_doctor_box.pack(padx=5, pady=5)

                ## Analys entry and labels ###

                analys_result_frame = Frame(blood_test_top, width=600, height=750)
                analys_result_frame.place(x=0, y=80)

                label_names = ["Creatine protein", "ASO Antistreptolysin-O", "WBC leukocyte count", "The amount of RBC erythrocytes", "The amount of HGB hemoglobin", "HCT hematocrit", "MCV mean volume of erythrocytes", "MCH 1 is the average volume of hemoglobin in erythrocytes", "MCHC is the average concentration of hemoglobin in erythrocytes", "The amount of PLT platelets", "RDW_SD distribution width of erythrocytes", "RDW_CV distribution width of erythrocytes", "PDW distribution width of platelets", "MPV average volume of platelets", "Amount of P_LCR platelets in % 18.9", "The percentage of PCT platelets is 0.28", "NEUT# total neutrophil count", "LYM# total number of lymphocytes", "MONO# Total number of monocytes", "EO# Total number of eosinophils", "BASO# total number of basophils", "NEUT% is the percentage of neutrophils", "LYM% is the amount of lymphocytes in %", "MONO% amount of monocytes in %", "EO% is the amount of eosinophils in %", "BASO% is the percentage of basophils", "EHS <<>> method"]

                data_entries_List = []

                def SubmitButton():

                    entry_list = []

                    for entries in data_entries_List: 
                        if str(entries.get()) == "":
                            messagebox.showerror(title="Error", message="All required fields must be filled out")
                            break
                        else:
                            entry_list.append(entries.get())
                            if len(entry_list) == 27:
                                entry_list.append(appointed_doctor_box.get())  
                                messagebox.showinfo(title="Data", message="Test results successfully exported as XLSX file to \"blood test results\" folder")
                                patients_blood_test_results.update({fin: entry_list})
                                try:
                                    results = json.load(open("blood_test_results.txt"))
                                except IOError:
                                    json.dump(patients_blood_test_results, open("blood_test_results.txt", "w"))
                                    # EXCEL EXPORTING  #
                                    ExportExcell()                               
                                else:
                                    patients_blood_test_results.update(results)
                                    json.dump(patients_blood_test_results, open("blood_test_results.txt", "w"))

                                    ExportExcell()

                row = 0  # Analizin adlarının label kimi göstərilməsi
                for k in label_names:
                    Label(analys_result_frame, text=f"{k}", font=("Microsoft YaHei UI", 10)).place(x=5, y= row)
                    row += 25

                row2 = 0
                for x in range(27): 
                    data_entry = Entry(analys_result_frame)
                    data_entry.place(x= 450, y = row2)
                    data_entries_List.append(data_entry)
                    row2 += 25
                
                submit_btn = Button(analys_result_frame, text="Submit", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=SubmitButton)
                submit_btn.place(x=500, y=680)
            else:
                messagebox.showerror(title="Searching error", message="Patient with this FIN was not found. Please register first.")

        else:
            if fin in results:

                name = patients_data[fin][0]
                surname = patients_data[fin][1]
                age = patients_data[fin][4]

                blood_test_top = Toplevel(edit_appointment_top)
                blood_test_top.title("Blood test data")
                blood_test_top.geometry("600x800+700+150")
                blood_test_top.resizable(False, False)

                search_frame = LabelFrame(blood_test_top, text="Patient's name and surname", font=("Microsoft YaHei UI", 11, "bold"))
                search_frame.place(x=10, y=10)

                name_lbl = Label(search_frame, text=f"{name} {surname}", font=("Microsoft YaHei UI", 13))
                name_lbl.pack(padx=5, pady=5)

                birth_data_frame = LabelFrame(blood_test_top, text="Age", font=("Microsoft YaHei UI", 11, "bold"))
                birth_data_frame.place(x=500, y=10)

                birth_data_lbl = Label(birth_data_frame, text=age, font=("Microsoft YaHei UI", 13))
                birth_data_lbl.pack(padx=5, pady=5)

                appointed_doctor_frame = LabelFrame(blood_test_top, text="Appointed doctor", font=("Microsoft YaHei UI", 11, "bold"))  ##########################
                appointed_doctor_frame.place(x=280, y=10)

                app_doctors = []
                for i in doctor_dataDict:
                    app_doctors.append((doctor_dataDict[i][0] + " " + doctor_dataDict[i][1]))
                    
                appointed_doctor_box = ttk.Combobox(appointed_doctor_frame, values=app_doctors, state="readonly")
                appointed_doctor_box.pack(padx=5, pady=5)
                index = 0
                for i in app_doctors:
                    if i == results[fin][27]:
                        appointed_doctor_box.current(index)
                    index += 1                                         

                # analys entry and labels ###        

                analys_result_frame = Frame(blood_test_top, width=600, height=750)
                analys_result_frame.place(x=0, y=80)

                label_names = ["Creatine protein", "ASO Antistreptolysin-O", "WBC leukocyte count", "The amount of RBC erythrocytes", "The amount of HGB hemoglobin", "HCT hematocrit", "MCV mean volume of erythrocytes", "MCH 1 is the average volume of hemoglobin in erythrocytes", "MCHC is the average concentration of hemoglobin in erythrocytes", "The amount of PLT platelets", "RDW_SD distribution width of erythrocytes", "RDW_CV distribution width of erythrocytes", "PDW distribution width of platelets", "MPV average volume of platelets", "Amount of P_LCR platelets in % 18.9", "The percentage of PCT platelets is 0.28", "NEUT# total neutrophil count", "LYM# total number of lymphocytes", "MONO# Total number of monocytes", "EO# Total number of eosinophils", "BASO# total number of basophils", "NEUT% is the percentage of neutrophils", "LYM% is the amount of lymphocytes in %", "MONO% amount of monocytes in %", "EO% is the amount of eosinophils in %", "BASO% is the percentage of basophils", "EHS <<>> method"]

                data_entries_List = []

                def SubmitButton():

                    entry_list = []

                    for entries in data_entries_List:  
                        if str(entries.get()) == "":
                            messagebox.showerror(title="Error", message="All required fields must be filled out")
                            break
                        else:
                            entry_list.append(entries.get())
                            if len(entry_list) == 27:
                                entry_list.append(appointed_doctor_box.get())
                                messagebox.showinfo(title="END", message="COMPLETED")
                                patients_blood_test_results.update({fin: entry_list})
                                json.dump(patients_blood_test_results, open("blood_test_results.txt", "w"))

                                ExportExcell()

                row = 0
                for k in label_names:
                    Label(analys_result_frame, text=f"{k}", font=("Microsoft YaHei UI", 10)).place(x=5, y= row)
                    row += 25
    
                row2 = 0
                data_index = 0
    
                for x in range(27):
                    data_entry = Entry(analys_result_frame)
                    data_entry.place(x= 450, y = row2)
                    data_entry.insert(0, results[fin][data_index])
                    data_entries_List.append(data_entry)
                    row2 += 25
                    data_index += 1
       
                submit_btn = Button(analys_result_frame, text="Submit", font=("Microsoft YaHei UI", 11, "bold"), fg="black", bd=1, bg="#58C3EB", activebackground='#C6EEFF', cursor="hand2", command=SubmitButton)
                submit_btn.place(x=500, y=680)
            else:
                messagebox.showerror(title="Searching error", message="Patient with this FIN was not found. Please register first.")
      
        def ExportExcell():
             wb = load_workbook(filename='blood_test_template.xlsx')   # excel faylının nümunəsi
             ws = wb.active # aktiv iş vərəqi
            
             results = json.load(open("blood_test_results.txt")) 
             ws['C3'] = name + " " + surname  # pasient adı
             ws['C5'] = results[fin][27]   # həkim adı
             ws['F3'] = age  
             ws['F5'] = datetime.datetime.now()  # qeydiyyat tarixi
             ws['D11'] = results[fin][0]
             ws['D13'] = results[fin][1] 
             data_index = 2   
             for i in range(16,41):     # 16-40 cı xanalara dataların göndərilməsi                    
                 ws[f'D{i}'] = results[fin][data_index]
                 data_index += 1 
             wb.save(f"blood test results/{fin}_patient_blood_test.xlsx")  # faylın save olunması

    tests_frame = Frame(edit_appointment_top, width=1000, height=500)
    tests_frame.place(x=0, y=0)

    blood_test_btn = Button(tests_frame, image=blood_test_img, bg="#79D4F6", activebackground="#C6EEFF", cursor="hand2", command=BloodTest)
    blood_test_btn.place(x=20, y=20)

doctors_edit_img = PhotoImage(file="buttons/doctor_edit.png")
doctors_edit_btn = Button(side_menu_frame, image=doctors_edit_img, bg="#79D4F6", activebackground="#C6EEFF", cursor="hand2", command=DoctorsEdit)
doctors_edit_btn.place(x=20, y=30)

patients_edit_img = PhotoImage(file="buttons/patients_edit.png")
patients_edit_btn = Button(side_menu_frame, image=patients_edit_img, bg="#79D4F6", activebackground="#C6EEFF", cursor="hand2", command=PatientsEdit)
patients_edit_btn.place(x=20, y=350)

appointment_edit_img = PhotoImage(file="buttons/appointment.png")
appointment_edit_btn = Button(side_menu_frame, image=appointment_edit_img, bg="#79D4F6", activebackground="#C6EEFF", cursor="hand2", command=AppointmentEdit)
appointment_edit_btn.place(x=20, y=670)

win.mainloop()